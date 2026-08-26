"""
Exportacion a Excel, mismo formato que el archivo viejo del cliente
(wter/AuctionsWebFiltrosBOE171121.xlsx), con una hoja mas:
- PROXIMA APERTURA / CELEBRANDOSE / CONCLUIDAS (las 3 del formato viejo)
- CANCELADAS: Cancelada/Suspendida - antes quedaban mezcladas dentro de
  CONCLUIDAS (que el formato viejo no distinguia como hoja aparte), pedido
  explicito del cliente separarlas.
- fila 1: "Index - Auctions.Web" (aca: "Index - Auctions Broker")
- fila 2: encabezados (24 columnas)
- fila 3+: datos
"""

from openpyxl import Workbook

import db

HEADERS = db.HEADERS

SHEETS_EXACTOS = {
    "PROXIMA APERTURA": {"Próxima apertura"},
    "CELEBRANDOSE": {"Celebrándose"},
    # "Concluida en Portal de Subastas" y "Finalizada por Autoridad Gestora"
    # (BOE) + "Concluida" (Seguridad Social) son las 3 formas que toma un
    # estado REALMENTE concluido. Antes esta hoja era "todo lo que no sea
    # Proxima apertura/Celebrandose", asi que Cancelada y Suspendida caian
    # aca tambien (mezcladas con subastas concluidas de verdad) - confirmado
    # por el cliente: de 304 lotes de Malaga en esta pestaña, 161 eran en
    # realidad Cancelada/Suspendida, no Concluida.
    "CONCLUIDAS": {"Concluida en Portal de Subastas", "Finalizada por Autoridad Gestora", "Concluida"},
}
# Cancelada/Suspendida (y cualquier estado futuro que no encaje en las 3
# hojas de arriba) van aca, para no perder filas del export como antes de
# que existiera esta hoja, sin mezclarlas con subastas concluidas.
HOJA_RESTO = "CANCELADAS"


def _euro(valor):
    if valor is None:
        return ""
    return f"{valor:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")


def _pct(a, b):
    if not a or not b:
        return ""
    return f"{(a / b * 100):,.2f} %".replace(",", "X").replace(".", ",").replace("X", ".")


def _fila_a_excel_row(row: dict):
    return [
        _pct(row["cantidad_reclamada"], row["valor_subasta"]),
        _pct(row["puja_minima"], row["valor_subasta"]),
        row["estado"],
        row["tipo_subasta"],
        row["tipo_bien"],
        row["id"],
        row["lotes"],
        row["provincia"],
        row["localidad"],
        row["direccion"],
        row["descripcion"],
        row["referencia_catastral"],
        row["marca"],
        row["modelo"],
        row["matricula"],
        _euro(row["cantidad_reclamada"]),
        _euro(row["valor_tasacion"]),
        _euro(row["valor_subasta"]),
        _euro(row["tramos_entre_pujas"]),
        _euro(row["puja_minima"]) if row["puja_minima"] is not None else "Sin puja mínima",
        _euro(row["importe_deposito"]),
        row["nombre"],
        row["fecha_inicio"],
        row["fecha_conclusion"],
    ]


def exportar(filas: list, path):
    """filas: lista de dicts (mismo shape que db.query_lotes()).
    path puede ser una ruta de archivo o un objeto tipo archivo (BytesIO);
    openpyxl acepta los dos en wb.save()."""
    wb = Workbook()
    wb.remove(wb.active)

    estados_conocidos = set()
    for nombre_hoja, estados in SHEETS_EXACTOS.items():
        estados_conocidos |= estados
        ws = wb.create_sheet(nombre_hoja)
        ws.append(["Index - Auctions Broker"])
        ws.append(HEADERS)
        for row in filas:
            if row["estado"] in estados:
                ws.append(_fila_a_excel_row(row))

    ws = wb.create_sheet(HOJA_RESTO)
    ws.append(["Index - Auctions Broker"])
    ws.append(HEADERS)
    for row in filas:
        if row["estado"] not in estados_conocidos:
            ws.append(_fila_a_excel_row(row))

    wb.save(path)
    return path


if __name__ == "__main__":
    rows = db.query_lotes()
    out = exportar(rows, "prueba_export.xlsx")
    print(f"Exportado {len(rows)} filas a {out}")
